from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl
from constants.globalConstants import *
import json

class Test_Sauce():

    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")
    
    def teardown_method(self):
        self.driver.quit()
    
    def getData():
        return [("1","1"),("abc","123"),("deneme","secret_sauce")]
    
    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data
    
    def readInvalidDataFromJson():
        with open("data/invalidLogin.json", 'r') as json_file:
            data = json.load(json_file)
        return [(username, password) for username, password in zip(data["username"], data["password"])]
    
    def waitForElementVisible(self,locator,timeout=10):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    def test_valid_login(self):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        locateLogo = self.waitForElementVisible((By.XPATH,locate_logo))
        testResult = locateLogo == self.waitForElementVisible((By.XPATH,test_result_id))
        print(f"Logo gorundu mu?: {testResult}")
        listOfItems = WebDriverWait(self.driver, 10).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, "inventory_item")))
        print(f"Saucedemo sitesinde şu an {len(listOfItems)} adet ürün bulunmaktadır.")

    def test_locked_out_user(self):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "locked_out_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage.text == errorMessage1_text

    def test_empty_pass_login(self):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "")
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage.text == errorMessage2_text

    def test_empty_login(self):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        userNameInput.send_keys("")
        passwordInput.send_keys("")
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage.text == errorMessage3_text

    def test_add_to_cart(self):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        addtoCard = self.waitForElementVisible((By.XPATH,add_to_card_xpath))
        addtoCard.click()
        checkout = self.waitForElementVisible((By.XPATH, checkout_xpath))
        checkout.click()
        bikeLight = self.waitForElementVisible((By.XPATH, bike_light_product))
        assert bikeLight.text == bike_light_product_text
         
    def test_remove_from_card(self):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        addtoCard = self.waitForElementVisible((By.XPATH,add_to_card_xpath))
        addtoCard.click()
        checkout = self.waitForElementVisible((By.XPATH, checkout_xpath))
        checkout.click()
        removefromCard = self.waitForElementVisible((By.XPATH, remove_from_card))
        removefromCard.click()

    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_invalid_login(self,username,password):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage.text == errorMessage_text

    @pytest.mark.parametrize("username, password", readInvalidDataFromJson())
    def test_invalid_login(self,username,password):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage.text == errorMessage_text
        
    def test_empty_checkout_info(self):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        addtoCard = self.waitForElementVisible((By.XPATH,add_to_card_xpath))
        addtoCard.click()
        checkout = self.waitForElementVisible((By.XPATH, checkout_xpath))
        checkout.click()
        checkoutButton = self.waitForElementVisible((By.XPATH, checkout_button_xpath))
        checkoutButton.click()
        continueButton = self.waitForElementVisible((By.XPATH, continue_button_xpath))
        continueButton.click()
        errorMsg = self.waitForElementVisible((By.XPATH, checkout_error_xpath))
        assert errorMsg.text == errorMessage4_text
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl
from constants.globalConstants import *

class Test_Sauce():

    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")
    
    def teardown_method(self):
        self.driver.quit()
    
    def getData():
        return [("1","1"),("abc","123"),("deneme","secret_sauce")]
    
    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data

    def test_valid_login(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        locateLogo = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,locate_logo)))
        testResult = locateLogo == WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,test_result_id)))
        print(f"Logo gorundu mu?: {testResult}")
        listCountTest = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.CLASS_NAME,list_count_test)))
        print("Sayfa üzerindeki ürün sayisi:", len(listCountTest))

    def test_locked_out_user(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "locked_out_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        assert errorMessage.text == errorMessage1_text

    def test_empty_pass_login(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "")
        actions.perform()
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        assert errorMessage.text == errorMessage2_text

    def test_empty_login(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        userNameInput.send_keys("")
        passwordInput.send_keys("")
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        assert errorMessage.text == errorMessage3_text

    def test_add_to_cart(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        addtoCard = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,add_to_card_xpath)))
        addtoCard.click()
        checkout = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, checkout_xpath)))
        checkout.click()
        bikeLight = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, bike_light_product)))
        assert bikeLight.text == bike_light_product_text
         
    def test_remove_from_card(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        addtoCard = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,add_to_card_xpath)))
        addtoCard.click()
        checkout = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, checkout_xpath)))
        checkout.click()
        removefromCard = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, remove_from_card)))
        removefromCard.click()

    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_invalid_login(self,username,password):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform()
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        errorMessage =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        assert errorMessage.text == errorMessage_text

    def test_empty_checkout_info(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton.click()
        addtoCard = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,add_to_card_xpath)))
        addtoCard.click()
        checkout = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, checkout_xpath)))
        checkout.click()
        checkoutButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, checkout_button_xpath)))
        checkoutButton.click()
        continueButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, continue_button_xpath)))
        continueButton.click()
        errorMsg = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, checkout_error_xpath)))
        assert errorMsg.text == errorMessage4_text